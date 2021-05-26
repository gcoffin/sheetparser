"""uses openpyxl as the engine to read the Excel sheet."""

# color index: -8 and use
# http://blog.softartisans.com/2013/05/13/kb-excels-color-palette-explained/
from __future__ import print_function

import logging
import re

import openpyxl
import openpyxl.comments
import openpyxl.styles
import six

from ..documents import (BORDER_TOP, BORDER_LEFT,
                         BORDER_BOTTOM, BORDER_RIGHT,
                         CellRange, SheetDocument, WorkbookDocument)
from ..utils import EMPTY_CELL

SHEETSTATE_VISIBLE = openpyxl.worksheet.worksheet.Worksheet.SHEETSTATE_VISIBLE
logger = logging.getLogger('sheetparser')


def color_tuple(hexastring):
    if not re.match('[0-9A-F]{6}', hexastring):
        return None
    # can come with alpha channel: keep only last 3
    return tuple(int(x + y, 16) for x, y in
                 zip(hexastring[::2], hexastring[1::2]))[-3:]


# solution to the theme in wksheet_fmt.parent.loaded_theme.decode('utf-8')

class Fill(object):
    def __init__(self, cell):
        fill = cell.fill
        book = cell.parent.parent
        self.type = fill.tagname
        self.color1 = None
        self.color2 = None
        if self.type == 'patternFill':
            self.pattern = fill.patternType
            if self.pattern is not None:
                self.color1 = self.get_color(fill.fgColor, book)
                self.color2 = self.get_color(fill.bgColor, book)
        else:
            raise NotImplementedError()

    def get_color(self, color, book):
        if color.type == 'theme':
            return {'theme': color.theme}
        elif color.type == 'indexed':
            index = color.index
            if not isinstance(index, str):
                try:
                    index = book._colors[index]
                except IndexError:
                    return None
            return color_tuple(index)
        elif color.type == 'rgb':
            return color_tuple(color.rgb)
        else:
            raise ValueError('Unknown color type')

    def __repr__(self):
        return "<Fill %s %s %s %s>" % (self.type, self.pattern,
                                       self.color1, self.color2)


class Formatting(object):
    def __init__(self, cell):
        self._cell = cell
        self._border_mask = None
        self._fill = None

    @property
    def border_mask(self):
        if self._border_mask is None:
            border = self._cell.border
            if border is None:
                self._border_mask = 0
            else:
                self._border_mask = (
                        (BORDER_TOP * (border.top.style is not None)) |
                        (BORDER_LEFT * (border.left.style is not None)) |
                        (BORDER_BOTTOM * (border.bottom.style is not None)) |
                        (BORDER_RIGHT * (border.right.style is not None)))
        return self._border_mask

    @property
    def fill(self):
        if self._fill is None:
            self._fill = Fill(self._cell)
        return self._fill

    @property
    def is_filled(self):
        return self.fill.type != 'patternFill' or self.fill.pattern is not None


class opxlCell(object):
    def __init__(self, value, cell, wksheet_fmt, is_merged):
        self._cell = cell
        self.value = EMPTY_CELL if value is None else value
        self._wksheet_fmt = wksheet_fmt  # used to write back
        self._formatting = None
        self.is_merged = is_merged

    def get_cell(self):
        return self._cell

    @property
    def formatting(self):
        if self._formatting is None:
            self._formatting = Formatting(self._cell)
        return self._formatting

    def __repr__(self):
        return "<opxlCell %s %s>" % (self._cell.row, self._cell.column)

    @property
    def border_mask(self):
        return self.formatting.border_mask

    def has_borders(self, mask):
        return bool(self.border_mask & mask)

    @property
    def is_filled(self):
        return self.formatting.is_filled

    @property
    def fill(self):
        return self.formatting.fill

    @property
    def is_empty(self):
        return self.value == EMPTY_CELL

    def set_value(self, value):
        self.get_cell().value = value

    def set_style(self, style):
        self.get_cell().style = style

    def set_comment(self, text, author):
        cell = self.get_cell()
        comment = cell.comment
        if not comment:
            cell.comment = openpyxl.comments.Comment(text, author)
        else:
            comment.text = text
            comment.author = author


class EmptyCell(opxlCell):

    def __init__(self, column, row, sheet):
        self.column = column
        self.row = row
        super(EmptyCell, self).__init__(None, None, sheet, False)

    def get_cell(self):
        if self._cell is None:
            self._cell = self.sheet.cell(column=self.column, row=self.row)
        return self._cell

    def has_borders(self, mask):
        if self._cell is None:
            return False
        return super(EmptyCell).has_borders(mask)


class opxlExcelSheet(SheetDocument, CellRange):
    def __init__(self, wksheet_data, wksheet_fmt=None):
        self.name = wksheet_data.title
        self.wksheet_data = wksheet_data
        self.wksheet_fmt = wksheet_fmt
        self.merged = {}
        self.hidden_rows = {}
        if self.wksheet_fmt:
            for crange in self.wksheet_fmt.merged_cell_ranges:
                clo, rlo, chi, rhi = openpyxl.utils.range_boundaries(str(crange))
                for rowx in range(rlo, rhi + 1):
                    for colx in range(clo, chi + 1):
                        if (rlo, clo) != (rowx, colx):
                            self.merged[rowx, colx] = (rlo, clo)
        self.top, self.left = 1, 1  # wksheet.min_row, wksheet.min_column
        self.bottom = wksheet_data.max_row + 1
        self.right = wksheet_data.max_column + 1

    def is_hidden(self):
        return self.wksheet_data.sheet_state != SHEETSTATE_VISIBLE

    def is_hidden_row(self, rowidx):
        return self.wksheet_fmt.row_dimensions[rowidx + 1].hidden

    def cell(self, row, col):
        abs_row = self.top + row
        abs_col = self.left + col
        is_merged = False
        if (abs_row, abs_col) in self.merged:
            is_merged = True
            abs_row, abs_col = self.merged[abs_row, abs_col]
        try:
            return opxlCell(
                self.wksheet_data.cell(row=abs_row, column=abs_col).value,
                (self.wksheet_fmt.cell(row=abs_row, column=abs_col)
                 if self.wksheet_fmt else None),
                self.wksheet_fmt, is_merged)
        except IndexError:
            return EmptyCell(row, col, self.wksheet_fmt)

    def __repr__(self):
        return "<opxlExcelSheet %s>" % self.name


class opxlExcelWorkbook(WorkbookDocument):
    """A class to open workbooks and obtain sheets"""

    def __init__(self, filename, with_formatting=True):
        # I'd like to open it readonly but then the merged cells
        # are not loaded
        if with_formatting:
            self.wbk_fmt = openpyxl.load_workbook(filename=filename)
        else:
            self.wbk_fmt = None
        # and we need the data too because cell values are the formulas!!
        self.wbk_data = openpyxl.load_workbook(filename=filename, data_only=True)

    def __iter__(self):
        return (self[s] for s in self.wbk_data.sheetnames)

    def __getitem__(self, name_or_id):
        if isinstance(name_or_id, str):
            return opxlExcelSheet(self.wbk_data[name_or_id],
                                  self.wbk_fmt[name_or_id]
                                  if self.wbk_fmt else None)
        else:
            return opxlExcelSheet(self.wbk_data.worksheets[id],
                                  self.wbk_fmt.worksheets[id]
                                  if self.wbk_fmt else None)


load_workbook = opxlExcelWorkbook
