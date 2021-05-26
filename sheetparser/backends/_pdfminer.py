from __future__ import division
from __future__ import print_function

import bisect
import functools
import logging
import re
import sys

import six
from pdfminer.converter import TextConverter
from pdfminer.layout import (LAParams, LTContainer,
                             LTFigure, LTTextLineHorizontal)
from pdfminer.pdfdocument import PDFDocument, PDFTextExtractionNotAllowed
# pdfminer or pdfminer.six
from pdfminer.pdfinterp import PDFResourceManager, PDFPageInterpreter
from pdfminer.pdfpage import PDFPage
from pdfminer.pdfparser import PDFParser

from . import _array

MARGIN = 1
MIN_INTERSECT = 4
PAD = ' '
CHAR_SIZE = 4.7
RE_LONG_SPACES = re.compile(' {2}')


@functools.total_ordering
class Interval(object):
    __slots__ = ['a', 'b', 'margin']

    def __init__(self, a, b, margin=MARGIN):
        self.a = a
        self.b = b
        self.margin = margin

    def __or__(self, o):
        if not (o & self):
            return None
        return Interval(min(self.a, o.a), max(self.b, o.b))

    def __and__(self, o):
        a1 = max(self.a, o.a)
        b1 = min(self.b, o.b)
        if b1 > a1:
            return Interval(a1, b1)
        else:
            return None

    def hull(self, o):
        return Interval(min(self.a, o.a), max(self.b, o.b))

    def dist(self, o):
        x, y = sorted([self, o])
        return max(0, y.a - x.b)

    @property
    def middle(self):
        return (self.a + self.b) / 2

    @property
    def size(self):
        return self.b - self.a

    def __contains__(self, o):
        return self.b + self.margin >= o.b >= o.a >= self.a - self.margin

    def __gt__(self, o):
        return self.a > o.b

    def __eq__(self, o):
        return (self & o) is not None

    def __str__(self):
        return "[%s - %s]" % (self.a, self.b)

    def __repr__(self):
        return "<Interval(%s, %s)>" % (self.a, self.b)


class Position:
    def __init__(self, x0, y0, x1, y1):
        self.x0, self.y0, self.x1, self.y1 = x0, y0, x1, y1
        self.h = self.y1 - self.y0
        self.w = self.x1 - self.x0
        self.y_int = Interval(y0, y1, MARGIN * self.h)
        self.x_int = Interval(x0, x1, MARGIN * self.w)

    def __str__(self):
        return " x, y = ( %.0f, %.0f ) w, h = ( %.0f, %.0f )" % (
            self.x0, self.y0, self.w, self.h)

    @property
    def rightx(self):
        return self.x1

    def label(self):
        return "(%.0f, %.0f)" % (self.x0, self.y0)

    def center(self):
        return self.x0 + self.w / 2, self.y0 + self.h / 2

    def same_y(self, o):
        return o.y_int in self.y_int or self.y_int in o.y_int


@functools.total_ordering
class TextFrame(object):
    char_size = CHAR_SIZE

    def __init__(self, position, text):
        self.position = position
        self.text = text

    def split_vertical(self):
        lines = self.text.split('\n')
        line_height = self.position.h / len(lines)
        for i, t in enumerate(lines):
            tpos = Position(self.position.x0,
                            self.position.y1 - (i + 1) * line_height,
                            self.position.x1,
                            self.position.y1 - i * line_height)
            pad = int(self.position.w / self.char_size) - len(t)
            if pad < 0:
                pad = 0
            yield TextFrame(tpos, t + PAD * pad)

    def split_horizontal(self):
        if not self.text:
            yield self.text
            return
        char_size = self.position.w / len(self.text)
        x0 = self.position.x0
        sub_text = iter(RE_LONG_SPACES.split(self.text))
        try:
            for t in sub_text:
                x1 = x0 + char_size * len(t)
                tpos = Position(x0, self.position.y0,
                                self.position.x1, self.position.y1)
                yield TextFrame(tpos, t)
                x0 = x1 + char_size * len(next(sub_text))  # for spaces
        except StopIteration:
            pass

    def __eq__(self, o):
        return self.position.x0 == o.position.x0

    def __gt__(self, o):
        return self.position.x0 > o.position.x0

    def __repr__(self):
        return "<TF %s: %s>" % (self.position, repr(self.text))


class AlignedTextFrame(object):
    def __init__(self, y_int, frames=[]):
        self.interval = y_int
        self.frames = []
        self.frames.extend(frames)

    def __repr__(self):
        return repr(u"%s: %s" % (self.interval,
                                 ", ".join(i.text for i in self.frames)))

    def is_aligned(self, text_frame):
        intersect = text_frame.position.y_int & self.interval
        if intersect is None:
            return 0
        return intersect.size

    def add_frame(self, text_frame):
        self.frames.extend(text_frame.split_horizontal())
        self.frames.sort()

    def merge(self, aligned):
        new_interval = self.interval | aligned.interval
        if new_interval:
            self.interval = new_interval
            self.frames.extend(aligned.frames)
            self.frames.sort()

    def get_padded_texts(self):
        prev_pos_x = 0
        result = []
        for tf in sorted(self.frames):
            result.append(PAD * int((tf.position.x0 - prev_pos_x) / tf.char_size))
            prev_pos_x = tf.position.x1
            result.append(tf.text)
        return ''.join(result)


class Page(object):
    def __init__(self):
        self.aligned_frames = []

    def add_frame(self, frame):
        if not frame.text.strip():
            return
        candidates = []
        insert = False
        if not self.aligned_frames:
            self.aligned_frames.append(
                AlignedTextFrame(frame.position.y_int, [frame]))
            return
        i = 0
        for i in range(len(self.aligned_frames)):
            line = self.aligned_frames[i]
            alignment = line.is_aligned(frame)
            if alignment:
                candidates.append((alignment, line))
            if frame.position.y_int > line.interval:
                insert = True
                break
        if not candidates:
            if insert:
                self.aligned_frames.insert(i, AlignedTextFrame(
                    frame.position.y_int, [frame]))
            else:
                self.aligned_frames.append(AlignedTextFrame(
                    frame.position.y_int, [frame]))
        else:
            _, line = sorted(candidates)[-1]
            line.add_frame(frame)

    def add_text(self, text_frame):
        for tf in text_frame.split_vertical():
            self.add_frame(tf)

    def setMediaBox(self, mediabox):
        self.x, self.y, self.w, self.h = mediabox


class PdfTable(object):
    """receives aligned frames and build a 2d table, based
    on the calculation of tables and columns."""

    def __init__(self, aligned_frames):
        self.aligned_frames = aligned_frames
        self._calculate_columns()  # build a list of columns
        self._calculate_rows()  # check for missing lines

    def _calculate_columns(self):
        if self.aligned_frames:
            self.columns = sorted(i.position.x_int
                                  for i in self.aligned_frames[0].frames if i)
        else:
            self.columns = []
        for row in self.aligned_frames[1:]:
            self._add_row(row.frames)

    def _add_row(self, row):
        for cell in row:
            # what are the columns that intersect this cell
            ids = [i for i, col in enumerate(self.columns)
                   if col & cell.position.x_int]
            if len(ids) == 0:  # there's none: new columnn
                bisect.insort(self.columns, cell.position.x_int)
            else:  # there's at least one: restrict existing
                for i in ids:
                    self.columns[i] &= cell.position.x_int

    def _calculate_rows(self):
        #  the the middle y of aligned frames
        centers = [aligned_frame.interval.middle
                   for aligned_frame in self.aligned_frames if aligned_frame]
        if len(centers) < 3:
            return centers
        #  median height between consecutive centers
        candidate = sorted(i - j for i, j in zip(centers[:-1], centers[1:])
                           )[len(centers) // 2]
        _aligned_frames = []
        prev = None
        for aligned_frame in self.aligned_frames:
            if aligned_frame is None:
                continue
            if prev is not None:
                # merge this frame with previous if it's too close
                while aligned_frame.interval.b < prev - 1.5 * candidate:
                    _aligned_frames.append(None)
                    prev = prev - candidate
            _aligned_frames.append(aligned_frame)
            prev = aligned_frame.interval.middle
        self.aligned_frames = _aligned_frames

    def get_table(self):
        table = []
        for aligned_frame in self.aligned_frames:
            row = [''] * len(self.columns)
            if aligned_frame is not None:
                for frame in aligned_frame.frames:
                    for i, col in enumerate(self.columns):
                        if col & frame.position.x_int:
                            row[i] += str(frame.text)
                            break
            table.append(row)
        return table

    def merge_margin(self, small_col=15, margin=5):
        result = []
        for col in self.columns:
            if result:
                pre = result[-1]
                if pre.dist(col) < margin:
                    if min(col.size, pre.size) < small_col:
                        result.pop(-1)
                        col = pre.hull(col)
            result.append(col)
        self.columns = result
        return self


class TextAnalyzer(TextConverter):
    def __init__(self, *args, **kwargs):
        TextConverter.__init__(self, *args, **kwargs)
        self.pages = {}

    def render(self, item):
        if isinstance(item, LTFigure):
            return
        if isinstance(item, LTContainer):
            for child in item:
                self.render(child)
        if isinstance(item, LTTextLineHorizontal):
            (self.pages.setdefault(self.pageno - 1, Page())
             .add_text(TextFrame(Position(item.x0, item.y0, item.x1, item.y1),
                                 item.get_text())))

    def begin_page(self, page, ctm):
        result = TextConverter.begin_page(self, page, ctm)
        self.pages.setdefault(self.pageno, Page()).setMediaBox(page.mediabox)
        return result

    def end_page(self, page):
        return TextConverter.end_page(self, page)

    def receive_layout(self, ltpage):
        if self.showpageno:
            self.write_text('Page %s\n' % ltpage.pageid)
        self.render(ltpage)

    def all_text(self):
        for pdfpage, page in sorted(self.pages.items()):
            yield [aligned_frame.get_padded_texts()
                   for aligned_frame in page.aligned_frames]

    def get_result(self):
        return {pageno: PdfTable(page.aligned_frames).get_table()
                for pageno, page in sorted(self.pages.items())}

    def handle_undefined_char(self, font, cid):
        """A hacky solution for
        https://stackoverflow.com/questions/34108647/why-character-id-160-is-not-recognised-as-unicode-in-pdfminer
        """
        if cid == 160:
            return ' '
        logging.info('undefined: %r, %r' % (font, cid))
        return '(cid:%d)' % cid


def read_pdf(fp, password='', *page_numbers):
    # Create a PDF parser object associated with the file object.
    parser = PDFParser(fp)
    # Create a PDF document object that stores the document structure.
    # Supply the password for initialization.
    document = PDFDocument(parser, password)
    # Check if the document allows text extraction. If not, abort.
    if not document.is_extractable:
        raise PDFTextExtractionNotAllowed
    rsrcmgr = PDFResourceManager(caching=True)
    laparams = LAParams()
    laparams.all_texts = False
    device = TextAnalyzer(rsrcmgr, sys.stdout, laparams=laparams)
    interpreter = PDFPageInterpreter(rsrcmgr, device)
    # Process each page contained in the document.
    for page in PDFPage.create_pages(document):
        interpreter.process_page(page)
    device.close()
    return device.get_result()


def load_workbook(fp, password='', with_formatting=False):
    assert not with_formatting
    if isinstance(fp, str):
        with open(fp, 'rb') as f:
            return _array.rawWorkbook(read_pdf(f, password))
    else:
        return _array.rawWorkbook(read_pdf(fp, password))


def pdf2excel(inputname, outputname):
    import openpyxl

    with open(inputname, 'rb') as f:
        wb = openpyxl.Workbook(write_only=True)
        try:
            pages = load_workbook(f)
            npages = [i.name for i in pages]
            for i in range(min(npages), max(npages) + 1):
                ws = wb.create_sheet()
                for row in pages[i].data:
                    ws.append(row)
        finally:
            wb.save(outputname)


if __name__ == '__main__':
    pdf2excel(sys.argv[1], sys.argv[2])
